from dataclasses import dataclass
from datetime import datetime
from typing import List, Dict
from openpyxl import load_workbook


class AldenoneXLSX:
    """
    Represents Aldenone Excel file
    """

    def __init__(self, filepath: str):
        self.filepath = filepath
        self.worksheet = None
        self.column_headers = None
        self._read_xlsx()
        self._get_column_headers()
        self.data = self.extract_data()

    def __repr__(self):
        string = ""
        for pole_data in self.data:
            string += f"{pole_data}" + "\n"
        return string

    def _read_xlsx(self) -> None:
        workbook = load_workbook(self.filepath, data_only=True, read_only=True)
        self.worksheet = workbook['Sheet1']

    def _get_column_headers(self) -> None:
        column_header_cell_values = self.worksheet[1]
        self.column_headers = [cell.value for cell in column_header_cell_values]

    def extract_data(self) -> List:
        pole_data = []
        for row in range(2, self.worksheet.max_row + 1):
            pole_data.append(self._extract_single_pole_data(row))

        return pole_data

    def _extract_single_pole_data(self, row: int) -> Dict:
        pole_data = {}
        for index, header in enumerate(self.column_headers):
            column_number = index + 1
            cell_value = self.worksheet.cell(row=row, column=column_number).value
            pole_data[header] = self._convert_to_string_if_date_time_obj(cell_value)

        return pole_data

    @staticmethod
    def _convert_to_string_if_date_time_obj(item) -> str:
        if isinstance(item, datetime):
            month = item.strftime("%-m")
            day = item.strftime("%-d")
            item = f'{month}-{day}'

        return item
