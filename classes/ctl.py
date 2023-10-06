from dataclasses import dataclass, field
from typing import List
from openpyxl import load_workbook
import os


class CTLXLSX:
    """
    Represents an CTL XLSX file
    """

    def __init__(self, filepath: str):
        self.pole_data_list = []
        self.template_filepath = self._get_template_mac_windows()
        self.filepath = filepath
        self.workbook = None
        self._load_template_xlsx()

    def set_pole_data_list(self, lst: List['CTLPole']):
        self.pole_data_list = lst

    def write_to_cell(self, value: any, row: int, column: int, sheet='Sheet1'):
        self.workbook[sheet].cell(row=row, column=column, value=value)

    def save_workbook(self):
        self.workbook.save(self.filepath)

    def _load_template_xlsx(self):
        self.workbook = load_workbook(self.template_filepath)

    @staticmethod
    def _get_template_mac_windows():
        # Get the directory of the current script
        current_directory = os.path.dirname(os.path.abspath(__file__))

        # Go up one directory to the project folder
        project_directory = os.path.dirname(current_directory)

        # Construct the full path to the Excel file for Windows and Mac
        excel_path = os.path.join(project_directory, 'templates', 'ctl_permit_template.xlsx')

        return excel_path


@dataclass
class CTLPole:

    pole_number: str
    ctl_pole_number: str
    latitude: str
    longitude: str
    pole_height_class: str
    span_length: str
    communication_make_ready: str
    utility_make_ready: str
    attachment_list: list = field(default_factory=list)
    midspan_list: list = field(default_factory=list)
