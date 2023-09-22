from dataclasses import dataclass, field
from typing import List, Dict
from classes.coordinates import Coordinates
from openpyxl import load_workbook


class NWEXLSX:
    """
    Represents an NWE XLSX file
    """

    def __init__(self, filepath: str):
        self.pole_data_list = []
        self.filepath = filepath
        self.worksheet = None
        self.column_headers = None
        self._read_xlsx()
        self._get_column_headers()
        self.make_ready_start_col = 12
        self.make_ready_end_col = 18

    def _read_xlsx(self):
        workbook = load_workbook(self.filepath)
        self.worksheet = workbook['Sheet1']

    def _get_column_headers(self):
        column_header_cell_values = self.worksheet[14]
        self.column_headers = [cell.value for cell in column_header_cell_values]

        column = 12  # L
        self.column_headers[column - 1] = self.worksheet.cell(row=15, column=column).value

        column = 13  # M
        main_header = self.worksheet.cell(row=15, column=column).value
        subheader_1 = self.worksheet.cell(row=16, column=column).value
        combined_main_header_and_subheader_1 = f"{main_header} {subheader_1}"
        self.column_headers[column - 1] = combined_main_header_and_subheader_1
        column = 14  # N
        subheader_2 = self.worksheet.cell(row=16, column=column).value
        combined_main_header_and_subheader_2 = f"{main_header} {subheader_2}"
        self.column_headers[column - 1] = combined_main_header_and_subheader_2

        column = 15  # O
        main_header = self.worksheet.cell(row=15, column=column).value
        subheader = self.worksheet.cell(row=16, column=column).value
        combined_main_header_and_subheader = f"{main_header} {subheader}"
        self.column_headers[column - 1] = combined_main_header_and_subheader

        column = 16  # P
        self.column_headers[column - 1] = self.worksheet.cell(row=15, column=column).value

        column = 17  # Q
        self.column_headers[column - 1] = self.worksheet.cell(row=15, column=column).value

        column = 18  # R
        self.column_headers[column - 1] = self.worksheet.cell(row=15, column=column).value

        self.column_headers = [header for header in self.column_headers if header is not None]

    def extract_data(self):
        pole_start_rows = self._extract_pole_start_rows()
        pole_data = []
        for current_index, start_row in enumerate(pole_start_rows):
            end_row = self._extract_pole_end_row(current_index, pole_start_rows)
            pole_data.append(self._extract_single_pole_data(start_row, end_row))

        return pole_data

    def _extract_pole_start_rows(self) -> List:
        pole_start_rows = []
        for row in range(17, self.worksheet.max_row + 1):
            cell_value = self.worksheet.cell(row=row, column=1).value

            if cell_value is not None:
                pole_start_rows.append(row)

        return pole_start_rows

    def _extract_pole_end_row(self, start_row_index: int, pole_start_rows: List) -> int:
        next_index = start_row_index + 1

        if next_index == len(pole_start_rows):
            end_row = self.worksheet.max_row
        else:
            end_row = pole_start_rows[next_index]

        return end_row

    def _extract_single_pole_data(self, start_row: int, end_row: int) -> Dict:
        pole_data = {}
        for index, header in enumerate(self.column_headers):
            column_number = index + 1
            is_make_ready_data = self.make_ready_start_col <= column_number <= self.make_ready_end_col
            if is_make_ready_data:
                cells_tuple = self.worksheet.iter_rows(min_row=start_row, max_row=end_row, min_col=column_number,
                                                       max_col=column_number)
                cells_lst = list(cells_tuple)
                remove_inner_tuple = [cell[0] for cell in cells_lst]
                cell_values_lst = [cell.value for cell in remove_inner_tuple]
                pole_data[header] = cell_values_lst
            else:
                pole_data[header] = self.worksheet.cell(row=start_row, column=index + 1).value

        return pole_data


@dataclass
class NWEPole:
    """
    Contains the pole data provided by the NWE XLSX file
    """

    pole_number: str
    coordinates: Coordinates
    attachment_list: list = field(default_factory=list)
